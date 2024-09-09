#!/usr/bin/env python
# coding: utf-8

# In[1]:


import openpyxl
import psycopg2
from datetime import datetime

# Define the path to your Excel file
excel_file_path = r'C:\Users\apranj\Downloads\weather_data.xlsx'
excel_file_path = excel_file_path.replace('\u202a', '').replace('\u202b', '')

# Load the workbook and select the active worksheet
wb = openpyxl.load_workbook(excel_file_path)
ws = wb.active

# Define the database connection parameters
db_params = {
    'database': 'postgres',  # Fill in your database name
    'user': 'postgres',
    'password': '1234',
    'host': 'localhost',
    'port': '5432'
}

# Connect to the PostgreSQL database
def create_postgres_connection(db_params):
    try:
        connection = psycopg2.connect(
            dbname=db_params['database'],
            user=db_params['user'],
            password=db_params['password'],
            host=db_params['host'],
            port=db_params['port']
        )
        print("PostgreSQL Database connection established successfully.")
        return connection
    except psycopg2.Error as e:
        print(f"Error connecting to PostgreSQL Database: {e}")
        return None

conn = create_postgres_connection(db_params)
if conn is None:
    raise Exception("Unable to connect to the database.")

cur = conn.cursor()

# PostgreSQL Functions
def table_exists(connection, table_name):
    try:
        with connection.cursor() as cursor:
            cursor.execute("SELECT EXISTS (SELECT 1 FROM information_schema.tables WHERE table_name = %s)", (table_name,))
            table_exists = cursor.fetchone()[0]
        return table_exists
    except psycopg2.Error as e:
        print(f"Error checking if table exists: {e}")
        return False

def create_table(connection, table_name, columns, primary_key):
    try:
        with connection.cursor() as cursor:
            columns_def = ', '.join([f'"{col}" VARCHAR(255)' for col in columns])
            primary_key_def = f", PRIMARY KEY ({primary_key})"
            create_table_sql = f"CREATE TABLE {table_name} ({columns_def}{primary_key_def})"
            cursor.execute(create_table_sql)
            connection.commit()
            print(f"Table '{table_name}' created successfully.")
    except psycopg2.Error as e:
        print(f"Error creating table: {e}")

def get_existing_columns(connection, table_name):
    try:
        with connection.cursor() as cursor:
            cursor.execute(f"SELECT column_name FROM information_schema.columns WHERE table_name = %s", (table_name,))
            existing_columns = [row[0] for row in cursor.fetchall()]
        return existing_columns
    except psycopg2.Error as e:
        print(f"Error retrieving existing columns: {e}")
        return []

def alter_table(connection, table_name, columns):
    try:
        with connection.cursor() as cursor:
            existing_columns = get_existing_columns(connection, table_name)
            for column in columns:
                if column.lower() not in [col.lower() for col in existing_columns]:
                    alter_sql = f"ALTER TABLE {table_name} ADD COLUMN {column} VARCHAR(255)"
                    cursor.execute(alter_sql)
                    print(f"Altered table '{table_name}' to add column '{column}'.")
            connection.commit()
    except psycopg2.Error as e:
        print(f"Error altering table: {e}")
        connection.rollback()

def fetch_existing_data(connection, table_name, primary_key):
    try:
        with connection.cursor() as cursor:
            cursor.execute(f"SELECT * FROM {table_name} ORDER BY {primary_key}")
            rows = cursor.fetchall()
            columns = [desc[0] for desc in cursor.description]
        data_dict = {row[columns.index(primary_key)]: row for row in rows}
        return data_dict, columns
    except psycopg2.Error as e:
        print(f"Error fetching existing data: {e}")
        return {}, []

def update_data_in_postgres(connection, table_name, columns, data, primary_key, exclude_columns=['ingestion_timestamp']):
    try:
        with connection.cursor() as cursor:
            for row in data:
                key = row[columns.index(primary_key)]
                
                # Select the current row from the database
                select_sql = f'SELECT {", ".join(columns)} FROM {table_name} WHERE {primary_key} = %s'
                cursor.execute(select_sql, (key,))
                current_row = cursor.fetchone()
                
                # Check if the current row exists
                if current_row is None:
                    print(f"Record with {primary_key} = {key} not found.")
                    continue
                
                # Compare each column value
                update_needed = False
                update_set = []
                update_values = []
                for i, col in enumerate(columns):
                    if col != primary_key and col not in exclude_columns:
                        db_value = str(current_row[i])
                        new_value = str(row[columns.index(col)])
                        if db_value != new_value:
                            update_needed = True
                            update_set.append(f'"{col}" = %s')
                            update_values.append(new_value)
                
                if update_needed:
                    update_sql = f'UPDATE {table_name} SET {", ".join(update_set)} WHERE {primary_key} = %s'
                    cursor.execute(update_sql, update_values + [key])
                    
            connection.commit()
            print("Data updated successfully.")
    except psycopg2.Error as e:
        print(f"Error updating data: {e}")
        connection.rollback()

def insert_data_into_postgres(connection, table_name, columns, data, primary_key):
    try:
        with connection.cursor() as cursor:
            insert_sql = f'INSERT INTO {table_name} ({", ".join(columns)}) VALUES ({", ".join(["%s"] * len(columns))})'
            cursor.executemany(insert_sql, data)
            connection.commit()
            print("Data inserted successfully.")
    except psycopg2.Error as e:
        print(f"Error inserting data: {e}")
        connection.rollback()

def load_excel_data_into_postgres(db_params, table_name, excel_file_path, primary_key):
    try:
        wb = openpyxl.load_workbook(excel_file_path)
        ws = wb.active
        print("Excel file read successfully.")

        columns = [cell.value.lower().replace(' ', '_') for cell in ws[1]]
        # Exclude any dynamic timestamp columns from the database schema
        columns = [col for col in columns if not col.startswith('2024-')]
        data = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            data.append(row)

        connection = create_postgres_connection(db_params)
        if connection is None:
            return

        if not table_exists(connection, table_name):
            create_table(connection, table_name, columns, primary_key)
        else:
            alter_table(connection, table_name, columns)

        existing_data, existing_columns = fetch_existing_data(connection, table_name, primary_key)
        if existing_data:
            update_data_in_postgres(connection, table_name, columns, data, primary_key)
        else:
            insert_data_into_postgres(connection, table_name, columns, data, primary_key)

        connection.close()
    except Exception as e:
        print(f"Error loading Excel data into PostgreSQL: {e}")

# Update Excel File with Timestamp
def update_excel_with_timestamp(file_path, updated_rows):
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook.active
    timestamp_header = "Ingestion Timestamp"
    timestamp_column = None

    # Find or create the timestamp column
    for col in range(1, sheet.max_column + 1):
        if sheet.cell(row=1, column=col).value == timestamp_header:
            timestamp_column = col
            break

    if timestamp_column is None:
        timestamp_column = sheet.max_column + 1
        sheet.cell(row=1, column=timestamp_column).value = timestamp_header

    # Update timestamp for only changed rows
    for row_index, _ in updated_rows:
        sheet.cell(row=row_index, column=timestamp_column).value = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    workbook.save(file_path)

# Fetch existing data from the database, ordered by 'city'
cur.execute("SELECT * FROM weather_data ORDER BY city;")
db_data = cur.fetchall()

# Convert database data to a dictionary for quick lookup
column_names = [desc[0].lower() for desc in cur.description]
db_data_dict = {row[0]: row for row in db_data}

# Get Excel column names
excel_headers = [cell.value for cell in ws[1]]
excel_headers_normalized = [header.strip().lower() for header in excel_headers]

# Determine the index of the unique identifier column in Excel (assuming 'city' is the unique identifier)
unique_col_index = excel_headers_normalized.index('city')

# Find common columns between Excel and database
common_columns = set(excel_headers_normalized) & set(column_names)

if not common_columns:
    raise ValueError("No common columns between Excel headers and database columns.")

# Determine which columns to update
update_columns = [col for col in column_names if col in common_columns and col != 'city']
update_indices = [excel_headers_normalized.index(col) for col in update_columns]

# Prepare the SQL update query
set_clause = ", ".join([f"{col} = %s" for col in update_columns])
update_query = f"UPDATE weather_data SET {set_clause}, updated_timestamp = %s WHERE city = %s"

# Iterate through the rows in the Excel file, starting from the second row
rows_to_update = []
for row_index, row in enumerate(ws.iter_rows(min_row=2, values_only=True)):
    city = row[unique_col_index]
    if city in db_data_dict:
        db_row = db_data_dict[city]
        update_values = [row[idx] for idx in update_indices]

        # Check for differences
        if any(str(row[idx]) != str(db_row[column_names.index(col)]) for col, idx in zip(update_columns, update_indices)):
            updated_timestamp = datetime.now()
            update_values.append(updated_timestamp)
            update_values.append(city)
            rows_to_update.append((row_index + 2, tuple(update_values)))
    else:
        print(f"City {city} not found in the database.")

# Update only the changed rows in the database
for _, row in rows_to_update:
    cur.execute(update_query, row)

# Commit the changes and close the connection
conn.commit()
cur.close()
conn.close()

print(f"Rows updated: {len(rows_to_update)}")

# Update Excel File with Timestamp
def update_excel_with_timestamp(file_path, updated_rows):
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook.active
    timestamp_header = "Ingestion Timestamp"
    timestamp_column = None

    # Find or create the timestamp column
    for col in range(1, sheet.max_column + 1):
        if sheet.cell(row=1, column=col).value == timestamp_header:
            timestamp_column = col
            break

    if timestamp_column is None:
        timestamp_column = sheet.max_column + 1
        sheet.cell(row=1, column=timestamp_column).value = timestamp_header

    # Update timestamp for only changed rows
    for row_index, _ in updated_rows:
        sheet.cell(row=row_index, column=timestamp_column).value = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    workbook.save(file_path)

# Run the data load and update process
table_name = 'weather_data2'
primary_key = 'city'

# Ensure that the Excel file is updated with the timestamps
update_excel_with_timestamp(excel_file_path, [(row_index, _) for row_index, _ in rows_to_update])


# In[ ]:




