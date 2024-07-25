#!/usr/bin/env python
# coding: utf-8

# In[12]:


import psycopg2
from openpyxl import load_workbook
import datetime  # Ensure datetime module is imported

def create_postgres_connection(db_params):
    try:
        connection = psycopg2.connect(
            dbname=db_params['database'],
            user=db_params['user'],
            password=db_params['password'],
            host=db_params['host'],
            port=db_params['port']
        )
        print("PostgreSQL Database connection established successfully.")
        return connection
    except psycopg2.Error as e:
        print(f"Error connecting to PostgreSQL Database: {e}")
        return None

def table_exists(connection, table_name):
    try:
        cursor = connection.cursor()
        cursor.execute(f"SELECT EXISTS (SELECT FROM information_schema.tables WHERE table_name = '{table_name.lower()}')")
        exists = cursor.fetchone()[0]
        cursor.close()
        return exists
    except psycopg2.DatabaseError as e:
        print(f"Error checking if table exists: {e}")
        return False

def get_table_columns(connection, table_name):
    try:
        cursor = connection.cursor()
        cursor.execute(f"""
            SELECT column_name
            FROM information_schema.columns
            WHERE table_name = '{table_name.lower()}'
        """)
        columns = [row[0].lower() for row in cursor.fetchall()]
        cursor.close()
        return columns
    except psycopg2.DatabaseError as e:
        print(f"Error fetching table columns: {e}")
        return []

def create_table(connection, table_name, columns):
    try:
        cursor = connection.cursor()
        columns_def = ', '.join([f'"{col.lower()}" VARCHAR(255)' for col in columns])
        cursor.execute(f"CREATE TABLE {table_name.lower()} ({columns_def})")
        connection.commit()
        cursor.close()
        print(f"Table '{table_name}' created successfully.")
    except psycopg2.DatabaseError as e:
        print(f"Error creating table: {e}")
        connection.rollback()

def insert_data(connection, table_name, columns, data):
    try:
        cursor = connection.cursor()
        placeholders = ', '.join(['%s'] * len(columns))
        insert_columns = ', '.join([f'"{col.lower()}"' for col in columns])
        insert_sql = f'INSERT INTO {table_name.lower()} ({insert_columns}) VALUES ({placeholders})'
        
        for row in data:
            primary_key_value = row[columns.index(primary_key)]

            # Check if the row exists in the database
            cursor.execute(f'SELECT * FROM {table_name.upper()} WHERE {primary_key.upper()} = :1', (primary_key_value,))
            existing_row = cursor.fetchone()
        
        cursor.executemany(insert_sql, data)
        connection.commit()
        print("Data inserted successfully.")
    except psycopg2.Error as e:
        print(f"Error inserting data: {e}")
        connection.rollback()

def upsert_data(connection, table_name, columns, data, primary_key):
    try:
        cursor = connection.cursor()

        # Prepare SQL statements for INSERT and UPDATE
        placeholders = ', '.join([f'%s' for _ in columns])
        insert_columns = ', '.join([f'"{col.lower()}"' for col in columns])
        update_columns = ', '.join([f'"{col.lower()}" = EXCLUDED."{col.lower()}"' for col in columns if col.lower() != primary_key.lower()])
        upsert_sql = f'''
            INSERT INTO {table_name.lower()} ({insert_columns}) 
            VALUES ({placeholders})
            ON CONFLICT ("{primary_key.lower()}") 
            DO UPDATE SET {update_columns}
        '''

        cursor.executemany(upsert_sql, data)
        connection.commit()
        print("Data upserted successfully.")
    except psycopg2.DatabaseError as e:
        print(f"Error upserting data: {e}")
        connection.rollback()

def load_excel_data_into_postgres(db_params, table_name, excel_file_path, primary_key):
    try:
        wb = load_workbook(excel_file_path)
        ws = wb.active
        print("Excel file read successfully.")

        # Extract column names
        columns = [cell.value.lower().replace(' ', '_') for cell in ws[1]]

        # Extract data rows
        data = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            data.append([str(cell) if isinstance(cell, (int, float, bool, datetime.datetime)) else cell for cell in row])
       
        print("Columns in the Excel file:", columns)
        print("First few rows of data:", data[:5])

        connection = create_postgres_connection(db_params)
                 
        if connection:
            if table_exists(connection, table_name):
                print(f"Table '{table_name}' already exists. Checking columns...")
                table_columns = get_table_columns(connection, table_name)
                print(f"Columns in the PostgreSQL table '{table_name}': {table_columns}")

                if set(columns) != set(table_columns):
                    print("Mismatch between Excel file columns and PostgreSQL table columns.")
                    print("Excel columns:", columns)
                    print("Table columns:", table_columns)
                    return
                
                print(f"Updating/inserting data into '{table_name}'...")
                upsert_data(connection, table_name, columns, data, primary_key)
            else:
                print(f"Table '{table_name}' does not exist. Creating and inserting data.")
                create_table(connection, table_name, columns)
                insert_data(connection, table_name, columns, data)
           
            connection.close()
            print("PostgreSQL Database connection closed.")
   
    except FileNotFoundError as e:
        print(f"Error: {e}")
    except KeyError as e:
        print(f"KeyError: {e}")
    except Exception as e:
        print(f"An unexpected error occurred: {e}")

# Database parameters for PostgreSQL
db_params = {
    'database': 'uploading_data',
    'user': 'postgres',
    'password': '1234',
    'host': 'localhost',
    'port': '5432'  # Default PostgreSQL port
}

# Define the table name and the path to the Excel file
table_name = 'weather_forecast'  # Adjust table name accordingly
excel_file_path = r'C:\Users\wissen\Downloads\weather_data.xlsx'
primary_key = 'city'  # Assuming 'city' is the primary key in lowercase

# Remove any unwanted Unicode characters
excel_file_path = excel_file_path.replace('\u202a', '').replace('\u202b', '')

# Call the function to create the table and ingest data
load_excel_data_into_postgres(db_params, table_name, excel_file_path, primary_key)

