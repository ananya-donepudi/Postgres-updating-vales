#!/usr/bin/env python
# coding: utf-8

# In[2]:


import psycopg2
from openpyxl import load_workbook
import datetime  # Ensure datetime module is imported

def create_postgres_connection(db_params):
    try:
        connection = psycopg2.connect(
            dbname=db_params['database'],
            user=db_params['user'],
            password=db_params['password'],
            host=db_params['host'],
            port=db_params['port']
        )
        print("PostgreSQL Database connection established successfully.")
        return connection
    except psycopg2.Error as e:
        print(f"Error connecting to PostgreSQL Database: {e}")
        return None

def table_exists(connection, table_name):
    try:
        cursor = connection.cursor()
        cursor.execute("SELECT EXISTS (SELECT 1 FROM information_schema.tables WHERE table_name = %s)", (table_name,))
        table_exists = cursor.fetchone()[0]
        cursor.close()
        return table_exists
    except psycopg2.Error as e:
        print(f"Error checking if table exists: {e}")
        return False

def create_table(connection, table_name, columns):
    try:
        cursor = connection.cursor()
        columns_def = ', '.join([f'"{col}" VARCHAR(255)' for col in columns])  # Adjust data types as needed
        create_table_sql = f"CREATE TABLE {table_name} ({columns_def})"
        cursor.execute(create_table_sql)
        cursor.close()
        connection.commit()
        print(f"Table '{table_name}' created successfully.")
    except psycopg2.Error as e:
        print(f"Error creating table: {e}")

def insert_data(connection, table_name, columns, data):
    try:
        cursor = connection.cursor()
        placeholders = ', '.join(['%s'] * len(columns))
        insert_columns = ', '.join([f'"{col}"' for col in columns])
        insert_sql = f'INSERT INTO {table_name} ({insert_columns}) VALUES ({placeholders})'
       
        cursor.executemany(insert_sql, data)
        connection.commit()
        print("Data inserted successfully.")
    except psycopg2.Error as e:
        print(f"Error inserting data: {e}")

def update_or_insert_data(connection, table_name, columns, data, primary_key):
    try:
        cursor = connection.cursor()

        # Prepare SQL statements
        update_columns = ', '.join([f'"{col}" = EXCLUDED."{col}"' for col in columns])
        insert_columns = ', '.join([f'"{col}"' for col in columns])
        placeholders = ', '.join(['%s'] * len(columns))

        # Convert datetime objects to strings
        data = [[str(cell) if isinstance(cell, datetime.datetime) else cell for cell in row] for row in data]

        insert_sql = f'INSERT INTO {table_name} ({insert_columns}) VALUES ({placeholders}) ON CONFLICT ("{primary_key}") DO UPDATE SET {update_columns}'

        cursor.executemany(insert_sql, data)
        connection.commit()
        print("Data updated/inserted successfully.")
    except psycopg2.Error as e:
        print(f"Error updating/inserting data: {e}")

def load_excel_data_into_postgres(db_params, table_name, excel_file_path, primary_key):
    try:
        wb = load_workbook(excel_file_path)
        ws = wb.active
        print("Excel file read successfully.")

        # Extract column names
        columns = [cell.value.lower().replace(' ', '_') for cell in ws[1]]

        # Extract data rows
        data = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            data.append([str(cell) if isinstance(cell, (int, float, bool, datetime.datetime)) else cell for cell in row])
       
        print("Columns in the Excel file:", columns)
        print("First few rows of data:", data[:5])

        connection = create_postgres_connection(db_params)
                 
        if connection:
            if table_exists(connection, table_name):
                print(f"Table '{table_name}' already exists. Updating/inserting data...")
                update_or_insert_data(connection, table_name, columns, data, primary_key)
            else:
                print(f"Table '{table_name}' does not exist. Creating and inserting data.")
                create_table(connection, table_name, columns)
                insert_data(connection, table_name, columns, data)
           
            connection.close()
            print("PostgreSQL Database connection closed.")
   
    except FileNotFoundError as e:
        print(f"Error: {e}")
    except KeyError as e:
        print(f"KeyError: {e}")
    except Exception as e:
        print(f"An unexpected error occurred: {e}")

# Database parameters for PostgreSQL
db_params = {
    'database': 'uploading_data',
    'user': 'postgres',
    'password': '1234',
    'host': 'localhost',
    'port': '5432'  # Default PostgreSQL port
}

# Define the table name and the path to the Excel file
table_name = 'dataset'  # Adjust table name accordingly
excel_file_path = r'C:\Users\wissen\Desktop\details.xlsx'
primary_key = 'id'  # Assuming 'id' is the primary key in lowercase

# Remove any unwanted Unicode characters
excel_file_path = excel_file_path.replace('\u202a', '').replace('\u202b', '')

# Call the function to create the table and ingest data
load_excel_data_into_postgres(db_params, table_name, excel_file_path, primary_key)


# In[ ]:




