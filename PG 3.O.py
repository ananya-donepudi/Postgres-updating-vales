#!/usr/bin/env python
# coding: utf-8

# In[22]:


import openpyxl
import psycopg2
from psycopg2 import sql
from datetime import datetime

# Function to create PostgreSQL connection
def create_postgres_connection(db_params):
    try:
        connection = psycopg2.connect(
            dbname=db_params['database'],
            user=db_params['user'],
            password=db_params['password'],
            host=db_params['host'],
            port=db_params['port']
        )
        print("PostgreSQL Database connection established successfully.")
        return connection
    except psycopg2.Error as e:
        print(f"Error connecting to PostgreSQL Database: {e}")
        return None

# Function to check if a table exists in the PostgreSQL database
def table_exists(connection, table_name):
    try:
        with connection.cursor() as cursor:
            cursor.execute("SELECT EXISTS (SELECT 1 FROM information_schema.tables WHERE table_name = %s)", (table_name,))
            return cursor.fetchone()[0]
    except psycopg2.Error as e:
        print(f"Error checking if table exists: {e}")
        return False

# Function to fetch existing data from the database
def fetch_existing_data(connection, table_name, primary_key):
    try:
        with connection.cursor() as cursor:
            cursor.execute(f"SELECT * FROM {table_name}")
            rows = cursor.fetchall()
            columns = [desc[0] for desc in cursor.description]
        data_dict = {row[columns.index(primary_key)]: row for row in rows}
        return data_dict, columns
    except psycopg2.Error as e:
        print(f"Error fetching existing data: {e}")
        return {}, []

# Function to update data in the PostgreSQL table
def update_data_in_postgres(connection, table_name, columns, data, primary_key):
    try:
        with connection.cursor() as cursor:
            for row in data:
                key = row[columns.index(primary_key)]
                update_set = []
                update_values = []
                
                for i, col in enumerate(columns):
                    if col != primary_key and col != 'ingestion_timestamp':
                        update_set.append(f'"{col}" = %s')
                        update_values.append(row[i])

                if update_set:
                    update_sql = f'UPDATE {table_name} SET {", ".join(update_set)}, ingestion_timestamp = %s WHERE {primary_key} = %s'
                    cursor.execute(update_sql, update_values + [datetime.now(), key])
               
            connection.commit()
            print("Data updated successfully.")
    except psycopg2.Error as e:
        print(f"Error updating data: {e}")
        connection.rollback()

# Function to update Excel with the ingestion timestamp
def update_excel_with_timestamp(file_path, updated_rows):
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook.active
    timestamp_header = "Ingestion Timestamp"
    timestamp_column = Nonex

    # Find or create the timestamp column
    for col in range(1, sheet.max_column + 1):
        if sheet.cell(row=1, column=col).value == timestamp_header:
            timestamp_column = col
            break

    if timestamp_column is None:
        timestamp_column = sheet.max_column + 1
        sheet.cell(row=1, column=timestamp_column).value = timestamp_header

    # Update the timestamp for only the rows that were updated
    for row_idx in updated_rows:
        sheet.cell(row=row_idx + 2, column=timestamp_column).value = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    workbook.save(file_path)

# Main function to load Excel data into PostgreSQL
def load_excel_data_into_postgres(db_params, table_name, excel_file_path, primary_key):
    try:
        wb = openpyxl.load_workbook(excel_file_path)
        ws = wb.active
        print("Excel file read successfully.")

        columns = [cell.value.lower().replace(' ', '_') for cell in ws[1]]
        data = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            row_data = [cell if isinstance(cell, (int, float, bool, datetime)) else str(cell) for cell in row]
            data.append(row_data)

        connection = create_postgres_connection(db_params)
       
        if connection:
            if table_exists(connection, table_name):
                print(f"Table '{table_name}' already exists. Fetching existing data...")
                existing_data, existing_columns = fetch_existing_data(connection, table_name, primary_key)
               
                rows_to_update = []
                updated_rows = []

                for idx, row in enumerate(data):
                    key = row[columns.index(primary_key)]
                    if key in existing_data:
                        existing_row = existing_data[key]
                        if any(str(row[columns.index(col)]) != str(existing_row[existing_columns.index(col)]) for col in columns if col != 'ingestion_timestamp'):
                            rows_to_update.append(row)
                            updated_rows.append(idx)

                # Update existing rows in the database
                if rows_to_update:
                    update_data_in_postgres(connection, table_name, columns, rows_to_update, primary_key)
                    update_excel_with_timestamp(excel_file_path, updated_rows)
       
            else:
                print(f"Table '{table_name}' does not exist. Creating and inserting data.")
                # Create table and insert data logic here

            connection.close()
            print("PostgreSQL Database connection closed.")
    except Exception as e:
        print(f"An unexpected error occurred: {e}")

# Define the database parameters, table name, and Excel file path
db_params = {
    'database': 'uploading_data',
    'user': 'postgres',
    'password': '1234',
    'host': 'localhost',
    'port': '5432'
}

table_name = 'weather_data'
excel_file_path = r'C:\Users\wissen\weather_data.xlsx'
primary_key = 'city'

# Load Excel data into PostgreSQL
load_excel_data_into_postgres(db_params, table_name, excel_file_path, primary_key)


# In[ ]:




